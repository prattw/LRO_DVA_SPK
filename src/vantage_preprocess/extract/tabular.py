"""
CSV and Excel (xlsx) ingestion: flatten tabular data to plain text for chunking / Vantage.

Uses stdlib :mod:`csv` and :mod:`openpyxl` (no pandas). Legacy ``.xls`` is not supported.
"""

from __future__ import annotations

import csv
import io
import logging
from io import BytesIO

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

_DEFAULT_MAX_ROWS = 100_000


def extract_csv_to_text(
    content: bytes,
    *,
    max_rows: int = _DEFAULT_MAX_ROWS,
) -> str:
    """
    Decode CSV/TSV bytes and render rows as tab-separated lines (one row per line).

    Encoding: UTF-8 with BOM handled via ``utf-8-sig``, then replacement for invalid bytes.
    """
    text = content.decode("utf-8-sig", errors="replace")
    buf = io.StringIO(text)
    try:
        sample = text[:4096]
        dialect = csv.Sniffer().sniff(sample, delimiters=",;\t|")
    except csv.Error:
        dialect = csv.excel
    buf.seek(0)
    reader = csv.reader(buf, dialect)
    lines: list[str] = []
    for i, row in enumerate(reader):
        if i >= max_rows:
            lines.append(f"... truncated after {max_rows} data rows ...")
            logger.warning("CSV row limit reached (%s rows)", max_rows)
            break
        lines.append("\t".join(cell or "" for cell in row))
    return "\n".join(lines)


def extract_xlsx_to_text(
    content: bytes,
    *,
    max_rows_per_sheet: int = _DEFAULT_MAX_ROWS,
) -> str:
    """
    Read all sheets from an ``.xlsx`` / ``.xlsm`` workbook; each sheet as a labeled section.

    Rows are tab-separated; ``read_only`` mode helps large files.
    """
    bio = BytesIO(content)
    wb = load_workbook(bio, read_only=True, data_only=True)
    parts: list[str] = []
    try:
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            parts.append(f"## Sheet: {sheet_name}")
            n = 0
            for row in ws.iter_rows(values_only=True):
                if n >= max_rows_per_sheet:
                    msg = (
                        f"... truncated ({max_rows_per_sheet} rows max) "
                        f"in sheet {sheet_name!r} ..."
                    )
                    parts.append(msg)
                    logger.warning(
                        "Excel row limit reached for sheet %s (%s rows)",
                        sheet_name,
                        max_rows_per_sheet,
                    )
                    break
                line = "\t".join("" if c is None else str(c) for c in row)
                if line.strip():
                    parts.append(line)
                n += 1
    finally:
        wb.close()
    return "\n".join(parts)
