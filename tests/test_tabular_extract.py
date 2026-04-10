"""CSV and Excel text extraction."""

from __future__ import annotations

from io import BytesIO

import openpyxl

from vantage_preprocess.extract.engine import extract_csv_document, extract_xlsx_document
from vantage_preprocess.extract.tabular import extract_csv_to_text, extract_xlsx_to_text
from vantage_preprocess.models.enums import DocumentType


def test_extract_csv_to_text_roundtrip() -> None:
    raw = "a,b,c\n1,2,3\n"
    text = extract_csv_to_text(raw.encode("utf-8"))
    assert "1\t2\t3" in text
    assert "a\tb\tc" in text


def test_extract_csv_document() -> None:
    doc = extract_csv_document(b"x,y\nfoo,bar\n", "t.csv", "/tmp/t.csv")
    assert doc.document_type == DocumentType.CSV
    assert "foo" in doc.pages[0].text


def test_extract_xlsx_to_text() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.append(["h1", "h2"])
    ws.append(["v1", "v2"])
    buf = BytesIO()
    wb.save(buf)
    text = extract_xlsx_to_text(buf.getvalue())
    assert "Sheet:" in text
    assert "h1" in text and "v2" in text


def test_extract_xlsx_document() -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws["A1"] = "only"
    buf = BytesIO()
    wb.save(buf)
    doc = extract_xlsx_document(buf.getvalue(), "t.xlsx", "/tmp/t.xlsx")
    assert doc.document_type == DocumentType.XLSX
    assert "only" in doc.pages[0].text
