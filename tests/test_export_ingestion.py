"""Army Vantage ingestion exporters (CSV / JSONL / XLSX)."""

from __future__ import annotations

import csv
import json
from datetime import UTC, datetime
from pathlib import Path

from vantage_preprocess.export.csv_exporter import CsvVantageExporter
from vantage_preprocess.export.ingestion_record import VantageIngestionRecord
from vantage_preprocess.export.jsonl_exporter import JsonlVantageExporter
from vantage_preprocess.export.preview import format_ingestion_preview
from vantage_preprocess.export.xlsx_exporter import XlsxVantageExporter
from vantage_preprocess.models.document import ExportRow
from vantage_preprocess.models.enums import DocumentType, ExtractMethod


def _sample_row(**kwargs: object) -> VantageIngestionRecord:
    base = dict(
        document_id="a" * 32,
        source_filename="doc.pdf",
        original_file_type="pdf",
        chunk_id=f"{'a' * 32}-chunk-0001",
        page_start=1,
        page_end=2,
        section_title="SECTION 1",
        document_type="unknown",
        extraction_method="parse",
        extraction_confidence=1.0,
        chunk_text="Hello world.",
        processing_timestamp=datetime(2026, 4, 10, 12, 0, 0, tzinfo=UTC),
    )
    base.update(kwargs)
    return VantageIngestionRecord.model_validate(base)


def test_csv_roundtrip_columns(tmp_path: Path) -> None:
    r = _sample_row()
    p = tmp_path / "out.csv"
    CsvVantageExporter().write([r], p)
    text = p.read_text(encoding="utf-8")
    assert "document_id" in text and "original_file_type" in text
    assert "chunk_quality_score" in text and "extraction_mode" in text
    with p.open(encoding="utf-8", newline="") as f:
        rows = list(csv.DictReader(f))
    assert len(rows) == 1
    assert rows[0]["document_id"] == r.document_id
    assert rows[0]["chunk_text"] == "Hello world."
    assert rows[0]["section_title"] == "SECTION 1"


def test_jsonl_utf8_and_one_line_per_record(tmp_path: Path) -> None:
    r = _sample_row(chunk_text='Quote " and unicode: résumé')
    p = tmp_path / "out.jsonl"
    JsonlVantageExporter().write([r], p)
    lines = p.read_text(encoding="utf-8").strip().split("\n")
    assert len(lines) == 1
    obj = json.loads(lines[0])
    assert obj["chunk_text"] == 'Quote " and unicode: résumé'
    assert obj["processing_timestamp"].startswith("2026-")


def test_excel_formula_prefix(tmp_path: Path) -> None:
    r = _sample_row(chunk_text="=1+1 cmd injection test")
    p = tmp_path / "out.xlsx"
    XlsxVantageExporter().write([r], p)
    from openpyxl import load_workbook

    wb = load_workbook(p)
    ws = wb.active
    assert ws is not None
    headers = [c.value for c in ws[1]]
    col = headers.index("chunk_text") + 1
    val = ws.cell(row=2, column=col).value
    assert str(val).startswith("'")


def test_preview_sample() -> None:
    s = format_ingestion_preview([_sample_row(), _sample_row(chunk_id="b-chunk-0002")], limit=2)
    assert "chunk_id" in s
    assert "b-chunk-0002" in s or "chunk" in s


def test_export_row_conversion_roundtrip(tmp_path: Path) -> None:
    from vantage_preprocess.export.csv_export import write_csv

    er = ExportRow(
        schema_version="3",
        document_id="x" * 32,
        source_filename="f.txt",
        page_start=1,
        page_end=1,
        section_title=None,
        chunk_id="x" * 32 + "-chunk-0001",
        chunk_text="body",
        document_type=DocumentType.TXT,
        confidence=0.9,
        extracted_method=ExtractMethod.PARSE,
        mime_type="text/plain",
        source_sha256="c" * 64,
    )
    p = tmp_path / "via.csv"
    write_csv([er], p, semantic_document_type="unknown")
    rows = p.read_text(encoding="utf-8")
    assert "original_file_type" in rows
    assert "txt" in rows.lower()
